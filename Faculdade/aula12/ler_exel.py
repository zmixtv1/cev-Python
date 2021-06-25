from openpyxl import load_workbook
from openpyxl import Workbook
'''

wb = load_workbook("C:\\abc\\Pasta1.xlsx")

ws = wb['Planilha1']

for linha in ws.rows:
 for celula in linha:
     print(str(celula.value).ljust(25), end="")
 print("")
'''

'''
wb = load_workbook("C:\\abc\\pasta1.xlsx")

ws = wb['Planilha1']

for linha in range(2, 6):
  produto = ws.cell(column=1, row=linha).value
  preco = ws.cell(column=2, row=linha).value
  quant = ws.cell(column=3, row=linha).value
  tot = round(preco * quant, 2)
  print("Produto:".rjust(12), produto)
  print("Preço: R$".rjust(12), preco)
  print("Quantidade:".rjust(12), quant)
  print("Total:".rjust(12), tot)
  print("-------------------------------------")
'''

'''
#Entregar
wb = load_workbook("C:\\abc\\alunos.xlsx")

ws = wb['Planilha1']

for linha in range(2, 7):
  aluno = ws.cell(column=1, row=linha).value
  nota1 = ws.cell(column=2, row=linha).value
  nota2 = ws.cell(column=3, row=linha).value
  media = ws.cell(column=4, row=linha).value = (nota1+nota2) / 2

  sit = ""
  if media < 7:
    sit = "Reprovado"
  elif media >= 7:
    sit = "Aprovado"
  else:
    sit = "Erro de cálculo"
  print("Aluno:".rjust(10), aluno)
  print("Prova-1:".rjust(10), nota1)
  print("Prova-2:".rjust(10), nota2)
  print("Média:".rjust(10), media, sit)
  print("-------------------------------------")
'''
#Mudou aqui
'''
wb = load_workbook("C:\\abc\\alunos3.xlsx")

ws = wb['Planilha1']
ws['D1'] = "Média"
ws["E1"] = "Situação"
ws.column_dimensions['E'].width = 10
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10

for linha in range(2, 9):

  aluno = ws.cell(column=1, row=linha).value
  nota1 = ws.cell(column=2, row=linha).value
  nota2 = ws.cell(column=3, row=linha).value
  media = (nota1 + nota2) / 2
  ws.cell(column=4, row=linha, value=media)


  sit = ""
  if media < 7:
    sit = "Reprovado"
  elif media >= 7:
    sit = "Aprovado"
  else:
    sit = "Erro de cálculo"
  sit = ws.cell(column=5, row=linha, value=sit)
  sit = ws.cell(column=5, row=linha).value

  print("Aluno:".rjust(10), aluno)
  print("Prova-1:".rjust(10), nota1)
  print("Prova-2:".rjust(10), nota2)
  print("Média:".rjust(10), media, sit)
  print("-------------------------------------")

  wb.save("C:\\abc\\alunos3.xlsx")
'''
#entregar

wb = Workbook()
ws = wb.active
ws.title = "Planilha com Produtos"
ws['A1'] = "Seq."
ws['B1'] = "Produtos"
ws['C1'] = "Valores"
ws['D1'] = "Quantidades"
ws['E1'] = "Total produto"
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15

totGeral = 0
linha = 2
produto = ""
while produto != 'FIM':
   produto = input("Informe o nome do Produto [FIM - Termina o programa]: ")

   if produto == 'FIM':
       continue

   valor = float(input("Qual o valor do produto: "))
   quantidade = float(input("Qual a quantidade do produto: "))
   total = valor * quantidade
   totGeral += total

   seq = str(linha-1)
   ws.cell(row=linha, column=1, value=seq)
   ws.cell(row=linha, column=2, value=produto)
   ws.cell(row=linha, column=3, value=valor)
   ws.cell(row=linha, column=4, value=quantidade)
   ws.cell(row=linha, column=5, value=total)

   linha += 1

ws.cell(row=linha, column=5, value=totGeral)
ws.cell(row=linha, column=4, value="Total = ")
wb.save("c:\\abc\\construção.xlsx")
print ("Planilha salva em c:/abc/construção.xlsx")



